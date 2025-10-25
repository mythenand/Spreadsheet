import io
import re
from typing import Optional, List, Tuple

import numpy as np
import pandas as pd
import streamlit as st

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Laser Sheet Outlier Highlighter (Iterative)", layout="wide")

# ---------- Helpers ----------

def read_excel_any(file, sheet_name=None) -> pd.DataFrame:
    """Read Excel (xls/xlsx) without assuming header; return as raw DataFrame (no header)."""
    name = file.name.lower()
    engine = "openpyxl" if name.endswith(".xlsx") else "xlrd"
    file.seek(0)
    return pd.read_excel(file, sheet_name=sheet_name, header=None, engine=engine)

def list_sheets(file) -> List[str]:
    name = file.name.lower()
    engine = "openpyxl" if name.endswith(".xlsx") else "xlrd"
    file.seek(0)
    xl = pd.ExcelFile(file, engine=engine)
    sheets = xl.sheet_names
    file.seek(0)
    return sheets

def normalize_header(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = s.strip().lower()
    s = re.sub(r"[\s_/()-]+", "", s)
    return s

def locate_max_corr_column(headers: List[str]) -> Optional[int]:
    norm_headers = [normalize_header(h) for h in headers]
    candidates = [
        "maximumcorrosionin","maxcorrosionin","maximumcorrosion",
        "maxcorrosion","maximumcorrosioninch","maximumcorrosioninches"
    ]
    for idx, nh in enumerate(norm_headers):
        if nh in candidates:
            return idx
    for idx, nh in enumerate(norm_headers):
        if "corrosion" in nh and ("max" in nh or "maximum" in nh):
            return idx
    return None

def mad(series: pd.Series) -> float:
    x = pd.to_numeric(series, errors="coerce").dropna()
    if x.empty:
        return np.nan
    med = x.median()
    return (x - med).abs().median()

def robust_z(series: pd.Series) -> pd.Series:
    x = pd.to_numeric(series, errors="coerce")
    med = x.median(skipna=True)
    mad_val = mad(x)
    if pd.isna(mad_val) or mad_val == 0:
        return pd.Series([np.nan]*len(x), index=x.index)
    # 1.4826 makes MAD consistent with std under normality
    return (x - med) / (1.4826 * mad_val)

def sigma_clip_mask(s: pd.Series, k: float, two_sided: bool, max_iter: int = 5) -> Tuple[pd.Series, List[int]]:
    """
    Iterative sigma-clipping.
    Returns final boolean mask and list with #new outliers per iteration.
    """
    x = pd.to_numeric(s, errors="coerce").copy()
    keep = x.notna().values  # start with all non-NaN candidates
    out_mask = np.zeros(len(x), dtype=bool)
    iters_added = []

    for _ in range(max_iter):
        vals = x[keep]
        if vals.size < 3:
            iters_added.append(0)
            break
        mu = np.nanmean(vals)
        sd = np.nanstd(vals, ddof=1)
        if np.isnan(sd) or sd == 0:
            iters_added.append(0)
            break
        z = (x - mu) / sd
        if two_sided:
            new_out = (np.abs(z) > k) & keep & ~out_mask
        else:
            new_out = (z > k) & keep & ~out_mask

        added = int(new_out.sum())
        iters_added.append(added)
        if added == 0:
            break

        out_mask = out_mask | new_out
        keep = keep & (~out_mask)

    return pd.Series(out_mask, index=s.index), iters_added

def robust_mask(s: pd.Series, k: float, two_sided: bool) -> pd.Series:
    rz = robust_z(s)
    if two_sided:
        return rz.abs() > k
    return rz > k

def build_output_workbook(raw_df: pd.DataFrame,
                          sheet_name: str,
                          full_mask: pd.Series) -> bytes:
    """
    Create an .xlsx in memory:
      - Row1-2: project lines (bold)
      - Row3: headers (bold)
      - Row4+: data rows
      - Highlight entire data row light-blue when full_mask is True
    full_mask has length == len(raw_df) with True on outlier rows (only rows >=3 considered).
    """
    n_rows, n_cols = raw_df.shape
    first_data_idx = 3

    bold_font = Font(bold=True)
    fill_outlier = PatternFill(fill_type="solid", fgColor="ADD8E6")  # light blue

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name if sheet_name else "Sheet1"

    # Write values
    for r in range(n_rows):
        for c in range(n_cols):
            ws.cell(row=r+1, column=c+1, value=raw_df.iat[r, c])

    # Bold 1..3
    for r in [1, 2, 3]:
        if r <= ws.max_row:
            for c in range(1, n_cols+1):
                ws.cell(row=r, column=c).font = bold_font

    # Highlight outliers (data rows only)
    if full_mask is not None and len(full_mask) == n_rows:
        for r in range(first_data_idx, n_rows):
            if bool(full_mask.iloc[r]):
                for c in range(1, n_cols+1):
                    ws.cell(row=r+1, column=c).fill = fill_outlier

    # Simple column width heuristic
    for c in range(1, n_cols+1):
        col_letter = get_column_letter(c)
        texts = [str(raw_df.iat[r-1, c-1]) if r-1 < n_rows else "" for r in range(1, min(n_rows, 200)+1)]
        max_len = min(max((len(t) for t in texts if t is not None), default=0), 40)
        ws.column_dimensions[col_letter].width = max(10, max_len + 2)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

# ---------- UI ----------

st.title("Laser Sheet — Outlier Highlighter (Iterative & Robust)")
st.write(
    "Upload a **single Excel file (XLS/XLSX)** where:\n"
    "- **Row 1–2**: project lines (bolded)\n"
    "- **Row 3**: header row (bolded)\n"
    "- **Row 4+**: data rows (variable length)\n\n"
    "Pick outlier method: **Iterative sigma-clipping** or **Median+MAD (robust)**. "
    "The app highlights entire outlier rows in **light blue**."
)

uploaded = st.file_uploader("Upload Excel (xls/xlsx)", type=["xls", "xlsx"])
if not uploaded:
    st.stop()

# Pick sheet
try:
    sheets = list_sheets(uploaded)
except Exception as e:
    st.error(f"Could not read workbook: {e}")
    st.stop()
sheet = st.selectbox("Sheet", sheets, index=0)

# Method + params
method = st.selectbox("Outlier method", ["Sigma-clipping (iterative)", "Median + MAD (robust)"], index=0)
side = st.selectbox("Side", ["High only (> +K)", "Two-sided (|value| > K)"], index=0)
k = st.slider("K (threshold)", 1.0, 4.0, 1.0, 0.1)
max_iter = st.number_input("Max iterations (for sigma-clipping)", min_value=1, max_value=15, value=5, step=1)

# Read raw sheet (no header)
try:
    df_raw = read_excel_any(uploaded, sheet_name=sheet)
except Exception as e:
    st.error(f"Failed reading sheet: {e}")
    st.stop()

n_rows, n_cols = df_raw.shape
if n_rows < 3:
    st.error("Sheet does not contain at least 3 rows (two project lines + header).")
    st.stop()

# Extract header row (#3, 0-based index 2) and data (index >=3)
headers = df_raw.iloc[2].tolist()
data = df_raw.iloc[3:].copy()
if not data.empty:
    data = data.dropna(how="all")
else:
    st.warning("No data rows found beneath the header (row 3).")
    data = pd.DataFrame(columns=headers)

# Align columns to headers length
if len(headers) != data.shape[1]:
    data = data.iloc[:, :len(headers)]
data.columns = headers

# Locate target column
col_idx = locate_max_corr_column(headers)
if col_idx is None:
    st.error("Could not locate the 'Maximum Corrosion (in)' column (case/spacing tolerant).")
    st.stop()
target_col = headers[col_idx]
st.info(f"Target column: **{target_col}**")

# Compute outlier mask on DATA rows only
two_sided = (side == "Two-sided (|value| > K)")
iter_log = []
if data.empty:
    data_mask = pd.Series([False]*0, index=data.index)
else:
    s = pd.to_numeric(data[target_col], errors="coerce")

    if method.startswith("Sigma"):
        data_mask, iter_log = sigma_clip_mask(s, k=float(k), two_sided=two_sided, max_iter=int(max_iter))
    else:
        data_mask = robust_mask(s, k=float(k), two_sided=two_sided)

# Expand to full sheet index for highlighting
full_mask = pd.Series([False]*len(df_raw), index=df_raw.index)
if len(data_mask) > 0:
    full_mask.iloc[3:3+len(data_mask)] = data_mask.values

# Preview
preview = data.copy()
if method.startswith("Sigma"):
    st.caption(f"Sigma-clipping iteration adds: {iter_log} (new outliers per pass)")
preview["__OUTLIER__"] = data_mask.values if len(data_mask)>0 else []
st.subheader("Preview (first 200 data rows)")
st.dataframe(preview.head(200), use_container_width=True)
st.success(f"Total outliers flagged: {int(preview['__OUTLIER__'].sum())}")

# Build output workbook (.xlsx) with bold header rows and row-level highlighting
xlsx_bytes = build_output_workbook(
    raw_df=df_raw,
    sheet_name=sheet,
    full_mask=full_mask
)

st.download_button(
    "Download highlighted Excel (.xlsx)",
    data=xlsx_bytes,
    file_name="Laser_Outliers_Highlighted.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
